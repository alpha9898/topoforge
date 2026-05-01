from pathlib import Path
import xml.etree.ElementTree as ET

from openpyxl import Workbook

from models import Cable, Device, Topology
from services.drawio_generator import generate_drawio_xml
from services.ai_parser import apply_ai_suggestions_to_parsed, build_ai_suggestions, enrich_topology_connections
from services.excel_parser import normalize_header, parse_file
from services.layout_engine import layout_topology
from services.topology_builder import build_topology, infer_device_type, normalize_port
from services.topology_completion import complete_standard_topology
from services.topology_corrections import DeviceAdd, DeviceUpdate, TopologyCorrections, apply_topology_corrections
from services.validator import validate_topology


def test_header_detection_and_excel_flow():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Connections"
    sheet.append(["Device", "Port", "Connected To", "Peer Port", "Management IP", "Cable Type"])
    sheet.append(["Firewall-1", "eth1", "SW1", "G1/0/1", "10.0.0.1", "ethernet"])
    sheet.append(["SW1", "G1/0/1", "Firewall-1", "eth1", "", "ethernet"])
    path = Path(__file__).with_name("_tmp_lld.xlsx")
    try:
        workbook.save(path)
        parsed = parse_file(path)
    finally:
        import os

        if os.path.exists(path):
            os.remove(path)
    topology = validate_topology(build_topology(parsed))

    assert normalize_header("Hostname / Label") == "hostname label"
    assert len(topology.devices) == 2
    assert len(topology.cables) == 2
    assert topology.cables[0].label == "Firewall-1 eth1 -> SW1 Gi1/0/1"


def test_normalizers_and_inference():
    assert normalize_port("Ethernet 1") == "eth1"
    assert normalize_port("G1/0/1") == "Gi1/0/1"
    assert normalize_port("mgmt") == "Mgmt"
    assert infer_device_type("gui1fwall01") == "firewall"
    assert infer_device_type("PDU-1") == "pdu"


def test_missing_port_warning_and_drawio_xml():
    parsed = {
        "raw_devices": [{"name": "Firewall-1"}, {"name": "SW1"}],
        "raw_connections": [
            {"sourceDevice": "Firewall-1", "sourcePort": "", "targetDevice": "SW1", "targetPort": "Gi1/0/1", "cableType": "ethernet"}
        ],
        "issues": [],
    }
    topology = layout_topology(validate_topology(build_topology(parsed)))
    xml = generate_drawio_xml(topology)

    assert any(issue.code == "missing_source_port" for issue in topology.issues)
    assert "<mxfile" in xml
    assert "device-firewall-1" in xml
    assert "shape=mxgraph.cisco.security.firewall" in xml
    assert "shape=mxgraph.rack.hpe_aruba.switches" in xml
    assert "Firewall-1 ? -&gt; SW1 Gi1/0/1" in xml
    assert "Gi1/0/1" in xml
    assert "TABLE 1: CABLE REFERENCE" in xml
    assert "TABLE 2: SWITCH / OOB PORT SUMMARY" in xml
    assert "LEGEND" in xml
    assert "A1" in xml
    assert "Firewall-1" in xml
    assert "Gi1/0/1" in xml


def test_device_corrections_change_remove_and_add_devices():
    parsed = {
        "raw_devices": [{"name": "BadDevice"}, {"name": "SW1"}],
        "raw_connections": [
            {"sourceDevice": "BadDevice", "sourcePort": "eth1", "targetDevice": "SW1", "targetPort": "Gi1/0/1", "cableType": "ethernet"}
        ],
        "issues": [],
    }
    topology = validate_topology(build_topology(parsed))
    corrected = apply_topology_corrections(
        topology,
        TopologyCorrections(
            device_updates=[DeviceUpdate(id="baddevice", type="firewall", name="Firewall-1")],
            removed_device_ids=["sw1"],
            added_devices=[DeviceAdd(name="SW2", type="switch")],
        ),
    )

    assert [device.name for device in corrected.devices] == ["Firewall-1", "SW2"]
    assert corrected.devices[0].type == "firewall"
    assert corrected.cables == []


def test_drawio_source_device_cable_colors_are_deterministic():
    parsed = {
        "raw_devices": [{"name": "Server-1"}, {"name": "Server-2"}, {"name": "SW1"}],
        "raw_connections": [
            {"sourceDevice": "Server-1", "sourcePort": "eth0", "targetDevice": "SW1", "targetPort": "Gi1/0/1", "cableType": "ethernet"},
            {"sourceDevice": "Server-1", "sourcePort": "eth1", "targetDevice": "SW1", "targetPort": "Gi1/0/2", "cableType": "ethernet"},
            {"sourceDevice": "SW1", "sourcePort": "Gi1/0/3", "targetDevice": "Server-2", "targetPort": "eth0", "cableType": "ethernet"},
        ],
        "issues": [],
    }
    topology = layout_topology(validate_topology(build_topology(parsed)))
    xml = generate_drawio_xml(topology)
    root = ET.fromstring(xml)
    cells = {cell.attrib["id"]: cell for cell in root.iter("mxCell") if "id" in cell.attrib}

    assert "TABLE 3: SOURCE CABLE COLORS" in xml
    assert "Server-1" in xml
    assert "Server-2" in xml
    assert "YELLOW" in xml
    assert "RED" in xml
    assert "strokeColor=#FDD835" in cells["cable-001"].attrib["style"]
    assert "strokeWidth=3" in cells["cable-001"].attrib["style"]
    assert "strokeColor=#FDD835" in cells["cable-002"].attrib["style"]
    assert "strokeColor=#E53935" in cells["cable-003"].attrib["style"]
    assert "RED Server-2" in xml


def test_ai_helper_maps_sample_aliases_and_ignores_false_connections(monkeypatch):
    monkeypatch.delenv("GEMINI_API_KEY", raising=False)
    workbook = Workbook()
    devices = workbook.active
    devices.title = "Devices"
    devices.append(["Device", "Device Type", "Management IP", "Zone"])
    devices.append(["gui1swtch01", "switch", "10.123.11.11", "Site-1-links"])
    devices.append(["gui1swtch02", "switch", "10.123.11.12", "Site-1-links"])
    devices.append(["gui1fwall01", "firewall", "10.123.11.7", "Site-1-links"])
    devices.append(["gui1fwall02", "firewall", "10.123.11.8", "Site-1-links"])

    connections = workbook.create_sheet("Connections")
    connections.append(["Device", "Port", "Connected To", "Peer Port", "Cable Type"])
    connections.append(["SW1", "49:1", "Firewall-1", "port9", "ethernet"])
    connections.append(["SW2", "49:2", "Firewall-2", "port10", "ethernet"])
    connections.append(["gui1vpsie01", "switch1", "firewall", "to-fw2-port12", "ethernet"])

    path = Path(__file__).with_name("_tmp_ai_aliases.xlsx")
    try:
        workbook.save(path)
        parsed = parse_file(path)
    finally:
        import os

        if os.path.exists(path):
            os.remove(path)
    suggestions = build_ai_suggestions(parsed)
    ai_parsed = apply_ai_suggestions_to_parsed(parsed, suggestions)
    topology = build_topology(ai_parsed)
    device_names = {device.name for device in topology.devices}

    assert suggestions["status"] == "gemini_api_key_missing"
    assert any(item["alias"] == "SW1" and item["canonical"] == "gui1swtch01" for item in suggestions["alias_suggestions"])
    assert any(item["alias"] == "Firewall-1" and item["canonical"] == "gui1fwall01" for item in suggestions["alias_suggestions"])
    assert "SW1" not in device_names
    assert "Firewall-1" not in device_names
    assert not any(cable.sourcePort == "switch1" and cable.targetDeviceId == "firewall" for cable in topology.cables)


def test_connection_enrichment_updates_roles_labels_and_descriptions(monkeypatch):
    monkeypatch.delenv("GEMINI_API_KEY", raising=False)
    parsed = {
        "raw_devices": [{"name": "gui1vpsie01"}, {"name": "gui1swtch01"}],
        "raw_connections": [
            {"sourceDevice": "gui1vpsie01", "sourcePort": "iDRAC", "targetDevice": "gui1swtch01", "targetPort": "41", "cableType": "ethernet"}
        ],
        "issues": [],
    }
    topology = build_topology(parsed)
    enriched = enrich_topology_connections(topology, {"source": "local_rules"})

    assert enriched.cables[0].connectionRole == "management"
    assert enriched.cables[0].cableType == "management"
    assert enriched.cables[0].description
    assert enriched.aiSuggestions
    assert enriched.aiSuggestions["connection_enrichment"][0]["cable_id"] == "cable-001"


def test_standard_topology_completion_adds_external_chain_and_oob():
    parsed = {
        "raw_devices": [{"name": "gui1fwall01"}, {"name": "gui1swtch01"}, {"name": "gui1vpsie01"}],
        "raw_connections": [
            {"sourceDevice": "gui1fwall01", "sourcePort": "eth1", "targetDevice": "gui1swtch01", "targetPort": "49", "cableType": "ethernet"}
        ],
        "issues": [],
    }
    topology = complete_standard_topology(build_topology(parsed))
    device_ids = {device.id for device in topology.devices}
    labels = {cable.label for cable in topology.cables}

    assert {"admin", "vpn-gateway", "internet", "isp-1", "isp-2", "oob-mgmt"}.issubset(device_ids)
    assert "Admin VPN -> VPN Gateway Admin" in labels
    assert "VPN Gateway Internet -> Internet VPN" in labels
    assert "Internet ISP-1 -> ISP-1 WAN" in labels
    assert any(cable.sourceDeviceId == "oob-mgmt" and cable.targetDeviceId == "gui1vpsie01" for cable in topology.cables)
    assert any(cable.sourceDeviceId == "isp-1" and cable.targetDeviceId == "gui1fwall01" for cable in topology.cables)


def test_standard_topology_completion_does_not_duplicate_on_second_pass():
    parsed = {
        "raw_devices": [{"name": "gui1fwall01"}, {"name": "gui1vpsie01"}],
        "raw_connections": [],
        "issues": [],
    }
    topology = complete_standard_topology(build_topology(parsed))
    first_count = len(topology.cables)
    topology = complete_standard_topology(topology)

    assert len(topology.cables) == first_count


def test_port_anchor_intelligence_assigns_role_based_sides():
    topology = layout_topology(
        Topology(
            devices=[
                Device(id="isp-1", name="ISP-1", type="isp_router"),
                Device(id="fw1", name="Firewall-1", type="firewall"),
                Device(id="fw2", name="Firewall-2", type="firewall"),
                Device(id="sw1", name="SW1", type="switch"),
                Device(id="server1", name="Server-1", type="server"),
                Device(id="oob-mgmt", name="OOB-MGMT", type="switch"),
                Device(id="pdu-1", name="PDU-1", type="pdu"),
            ],
            cables=[
                Cable(id="wan", sourceDeviceId="fw1", sourcePort="WAN1", targetDeviceId="isp-1", targetPort="LAN", cableType="wan", connectionRole="wan"),
                Cable(id="lan", sourceDeviceId="fw1", sourcePort="eth1", targetDeviceId="sw1", targetPort="Gi1/0/49", cableType="ethernet", connectionRole="lan"),
                Cable(id="ha", sourceDeviceId="fw1", sourcePort="HA1", targetDeviceId="fw2", targetPort="HA1", cableType="ethernet", connectionRole="ha"),
                Cable(id="server", sourceDeviceId="sw1", sourcePort="Gi1/0/1", targetDeviceId="server1", targetPort="eth0", cableType="ethernet", connectionRole="lan"),
                Cable(id="mgmt", sourceDeviceId="server1", sourcePort="iDRAC", targetDeviceId="oob-mgmt", targetPort="1", cableType="management", connectionRole="management"),
                Cable(id="power", sourceDeviceId="server1", sourcePort="PSU1", targetDeviceId="pdu-1", targetPort="A1", cableType="power", connectionRole="power"),
            ],
        )
    )
    cables = {cable.id: cable for cable in topology.cables}

    assert (cables["wan"].exitX, cables["wan"].exitY) == (0.5, 0.0)
    assert cables["wan"].entryY == 1.0
    assert (cables["lan"].exitX, cables["lan"].exitY) == (0.5, 1.0)
    assert cables["lan"].entryY == 0.0
    assert cables["ha"].exitX == 1.0
    assert cables["ha"].entryX == 1.0
    assert cables["server"].exitY == 1.0
    assert (cables["server"].entryX, cables["server"].entryY) == (0.5, 0.0)
    assert cables["mgmt"].exitX == 0.0
    assert cables["mgmt"].entryY == 0.0
    assert (cables["power"].exitX, cables["power"].exitY) == (0.5, 1.0)
    assert (cables["power"].entryX, cables["power"].entryY) == (0.5, 0.0)


def test_port_anchor_intelligence_spreads_same_side_switch_ports():
    topology = layout_topology(
        Topology(
            devices=[
                Device(id="sw1", name="SW1", type="switch"),
                Device(id="server1", name="Server-1", type="server"),
                Device(id="server2", name="Server-2", type="server"),
                Device(id="server3", name="Server-3", type="server"),
            ],
            cables=[
                Cable(id="cable-001", sourceDeviceId="sw1", sourcePort="Gi1/0/1", targetDeviceId="server1", targetPort="eth0", cableType="ethernet", connectionRole="lan"),
                Cable(id="cable-002", sourceDeviceId="sw1", sourcePort="Gi1/0/2", targetDeviceId="server2", targetPort="eth0", cableType="ethernet", connectionRole="lan"),
                Cable(id="cable-003", sourceDeviceId="sw1", sourcePort="Gi1/0/3", targetDeviceId="server3", targetPort="eth0", cableType="ethernet", connectionRole="lan"),
            ],
        )
    )

    assert [(cable.exitX, cable.exitY) for cable in topology.cables] == [(0.25, 1.0), (0.5, 1.0), (0.75, 1.0)]


def test_port_anchor_intelligence_is_written_to_drawio_xml():
    topology = layout_topology(
        Topology(
            devices=[Device(id="fw1", name="Firewall-1", type="firewall"), Device(id="sw1", name="SW1", type="switch")],
            cables=[
                Cable(id="cable-001", sourceDeviceId="fw1", sourcePort="eth1", targetDeviceId="sw1", targetPort="Gi1/0/49", cableType="ethernet", connectionRole="lan")
            ],
        )
    )
    xml = generate_drawio_xml(topology)
    root = ET.fromstring(xml)
    cable_cell = next(cell for cell in root.iter("mxCell") if cell.attrib.get("id") == "cable-001")

    assert "exitX=0.5;exitY=1.0;entryX=0.5;entryY=0.0;" in cable_cell.attrib["style"]


def test_collision_avoidance_increases_column_spacing_for_dense_rows():
    simple = layout_topology(
        Topology(
            devices=[Device(id="sw1", name="SW1", type="switch"), Device(id="sw2", name="SW2", type="switch")],
            cables=[Cable(id="simple", sourceDeviceId="sw1", sourcePort="1", targetDeviceId="sw2", targetPort="1", cableType="ethernet", connectionRole="lan")],
        )
    )
    dense_devices = [Device(id="sw1", name="SW1", type="switch"), Device(id="sw2", name="SW2", type="switch")]
    dense_devices.extend(Device(id=f"server{index}", name=f"Server-{index}", type="server") for index in range(1, 7))
    dense = layout_topology(
        Topology(
            devices=dense_devices,
            cables=[
                Cable(
                    id=f"dense-{index}",
                    sourceDeviceId="sw1",
                    sourcePort=f"Gi1/0/{index}",
                    targetDeviceId=f"server{index}",
                    targetPort="eth0",
                    cableType="ethernet",
                    connectionRole="lan",
                )
                for index in range(1, 7)
            ],
        )
    )
    simple_switches = sorted([device for device in simple.devices if device.type == "switch"], key=lambda item: item.x)
    dense_switches = sorted([device for device in dense.devices if device.type == "switch"], key=lambda item: item.x)

    assert dense_switches[1].x - dense_switches[0].x > simple_switches[1].x - simple_switches[0].x


def test_collision_avoidance_increases_row_spacing_for_dense_inter_row_links():
    simple = layout_topology(
        Topology(
            devices=[Device(id="fw1", name="Firewall-1", type="firewall"), Device(id="sw1", name="SW1", type="switch")],
            cables=[Cable(id="link-1", sourceDeviceId="fw1", sourcePort="eth1", targetDeviceId="sw1", targetPort="49", cableType="ethernet", connectionRole="lan")],
        )
    )
    dense = layout_topology(
        Topology(
            devices=[Device(id="fw1", name="Firewall-1", type="firewall"), Device(id="sw1", name="SW1", type="switch")],
            cables=[
                Cable(id=f"link-{index}", sourceDeviceId="fw1", sourcePort=f"eth{index}", targetDeviceId="sw1", targetPort=str(48 + index), cableType="ethernet", connectionRole="lan")
                for index in range(1, 7)
            ],
        )
    )
    simple_gap = next(device.y for device in simple.devices if device.id == "sw1") - next(device.y for device in simple.devices if device.id == "fw1")
    dense_gap = next(device.y for device in dense.devices if device.id == "sw1") - next(device.y for device in dense.devices if device.id == "fw1")

    assert dense_gap > simple_gap


def test_auto_layout_centers_high_degree_devices_and_keeps_pairs_adjacent():
    topology = layout_topology(
        Topology(
            devices=[
                Device(id="sw1", name="SW1", type="switch"),
                Device(id="sw2", name="SW2", type="switch"),
                Device(id="sw3", name="SW3", type="switch"),
                Device(id="server1", name="Server-1", type="server"),
                Device(id="server2", name="Server-2", type="server"),
                Device(id="server3", name="Server-3", type="server"),
            ],
            cables=[
                Cable(id=f"server-{index}", sourceDeviceId="sw2", sourcePort=f"Gi1/0/{index}", targetDeviceId=f"server{index}", targetPort="eth0", cableType="ethernet", connectionRole="lan")
                for index in range(1, 4)
            ],
        )
    )
    switches = sorted([device for device in topology.devices if device.type == "switch"], key=lambda item: item.x)

    assert switches[1].id == "sw2"

    firewall_topology = layout_topology(
        Topology(
            devices=[Device(id=f"fw{index}", name=f"Firewall-{index}", type="firewall") for index in range(1, 5)],
            cables=[],
        )
    )
    firewalls = [device.id for device in sorted(firewall_topology.devices, key=lambda item: item.x)]

    assert abs(firewalls.index("fw1") - firewalls.index("fw2")) == 1


def test_parallel_cables_receive_distinct_waypoints():
    topology = layout_topology(
        Topology(
            devices=[Device(id="fw1", name="Firewall-1", type="firewall"), Device(id="sw1", name="SW1", type="switch")],
            cables=[
                Cable(id=f"cable-00{index}", sourceDeviceId="fw1", sourcePort=f"eth{index}", targetDeviceId="sw1", targetPort=str(48 + index), cableType="ethernet", connectionRole="lan")
                for index in range(1, 5)
            ],
        )
    )
    xml = generate_drawio_xml(topology)
    root = ET.fromstring(xml)
    points = [
        (point.attrib["x"], point.attrib["y"])
        for cable_id in ["cable-001", "cable-002", "cable-003", "cable-004"]
        for cell in root.iter("mxCell")
        if cell.attrib.get("id") == cable_id
        for point in cell.iter("mxPoint")
        if "as" not in point.attrib
    ]

    assert len(set(points)) == 4


def test_cable_reference_includes_expanded_columns_and_vlan_values():
    parsed = {
        "raw_devices": [{"name": "Firewall-1"}, {"name": "SW1"}],
        "raw_connections": [
            {"sourceDevice": "Firewall-1", "sourcePort": "eth1", "targetDevice": "SW1", "targetPort": "Gi1/0/1", "cableType": "ethernet", "role": "VLAN 603"},
            {"sourceDevice": "Firewall-1", "sourcePort": "eth2", "targetDevice": "SW1", "targetPort": "Gi1/0/2", "cableType": "ethernet", "role": "lan"},
        ],
        "issues": [],
    }
    topology = layout_topology(validate_topology(build_topology(parsed)))
    xml = generate_drawio_xml(topology)
    root = ET.fromstring(xml)
    cells = {cell.attrib["id"]: cell for cell in root.iter("mxCell") if "id" in cell.attrib}

    assert "Source Device" in xml
    assert "Source Port" in xml
    assert "Destination Device" in xml
    assert "Destination Port" in xml
    assert "VLAN" in xml
    assert "VLAN 603" in xml
    assert cells["cable-reference-row-1-7"].attrib["value"] == "-"
