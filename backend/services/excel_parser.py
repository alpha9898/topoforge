from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from models import Issue

HEADER_ALIASES: dict[str, set[str]] = {
    "device_name": {"device", "device name", "hostname", "hostname label", "node", "equipment", "label"},
    "hostname": {"hostname", "host", "fqdn"},
    "device_type": {"type", "device type", "role", "platform"},
    "port": {"port", "interface", "int", "nic", "eth", "source port"},
    "connected_to": {"connected to", "peer", "remote device", "destination", "target device"},
    "connected_port": {"peer port", "remote port", "destination port", "target port", "connected port"},
    "mgmt_ip": {"ip", "management ip", "mgmt ip", "idrac ip", "address"},
    "service_ip": {"service ip"},
    "vlan": {"vlan", "vlan id", "network", "segment"},
    "zone": {"zone", "site", "dc", "datacenter"},
    "cable_type": {"cable", "media", "link type", "cable type"},
}


def normalize_header(value: Any) -> str:
    text = str(value or "").replace("\xa0", " ").strip().lower()
    text = re.sub(r"[/_-]+", " ", text)
    text = re.sub(r"[^a-z0-9 ]+", "", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_file(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _parse_rows(path.stem, _read_csv(path))
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise RuntimeError("openpyxl is required to parse Excel files") from exc

        raw_devices: list[dict[str, Any]] = []
        raw_connections: list[dict[str, Any]] = []
        issues: list[Issue] = []
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                parsed = _parse_rows(sheet.title, list(sheet.iter_rows(values_only=True)))
                raw_devices.extend(parsed["raw_devices"])
                raw_connections.extend(parsed["raw_connections"])
                issues.extend(parsed["issues"])
        finally:
            workbook.close()
        return {"raw_devices": raw_devices, "raw_connections": raw_connections, "issues": issues}
    return {
        "raw_devices": [],
        "raw_connections": [],
        "issues": [
            Issue(
                id="parse-unsupported-file",
                severity="error",
                code="unsupported_file",
                message=f"Unsupported file type: {suffix}",
            )
        ],
    }


def _read_csv(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return [row for row in csv.reader(handle)]


def _parse_rows(sheet_name: str, rows: list[tuple[Any, ...] | list[Any]]) -> dict[str, Any]:
    rows = [tuple(cell if cell is not None else "" for cell in row) for row in rows]
    header_index, field_map = _detect_header(rows)
    raw_devices: list[dict[str, Any]] = []
    raw_connections: list[dict[str, Any]] = []
    issues: list[Issue] = []

    if header_index is None:
        raw_connections.extend(_scan_freeform_connections(sheet_name, rows))
        return {"raw_devices": raw_devices, "raw_connections": raw_connections, "issues": issues}

    headers = [str(cell or "") for cell in rows[header_index]]
    data_rows = rows[header_index + 1 :]
    for row_offset, row in enumerate(data_rows, start=header_index + 2):
        if not any(str(cell).strip() for cell in row):
            continue

        device_name = _get(row, field_map.get("device_name"))
        hostname = _get(row, field_map.get("hostname"))
        mgmt_ip = _get(row, field_map.get("mgmt_ip"))
        service_ip = _get(row, field_map.get("service_ip"))
        device_type = _get(row, field_map.get("device_type"))
        zone = _get(row, field_map.get("zone")) or sheet_name

        if device_name:
            raw_devices.append(
                {
                    "name": device_name,
                    "hostname": hostname or device_name,
                    "type": device_type,
                    "mgmtIp": mgmt_ip or service_ip,
                    "zone": zone,
                    "source": f"{sheet_name}:{row_offset}",
                }
            )

        if device_name and field_map.get("connected_to") is not None:
            raw_connections.append(
                {
                    "sourceDevice": device_name,
                    "sourcePort": _get(row, field_map.get("port")),
                    "targetDevice": _get(row, field_map.get("connected_to")),
                    "targetPort": _get(row, field_map.get("connected_port")),
                    "cableType": _get(row, field_map.get("cable_type")) or "unknown",
                    "role": _get(row, field_map.get("vlan")) or "unknown",
                    "source": f"{sheet_name}:{row_offset}",
                }
            )

        if device_name:
            raw_connections.extend(_extract_header_connections(sheet_name, headers, row, device_name, row_offset))

        raw_connections.extend(_scan_freeform_connections(sheet_name, [row], row_offset))

    if not raw_devices and not raw_connections:
        issues.append(
            Issue(
                id=f"parse-empty-{_slug(sheet_name)}",
                severity="warning",
                code="empty_sheet",
                message=f"No topology rows were detected in sheet {sheet_name}.",
            )
        )
    return {"raw_devices": raw_devices, "raw_connections": raw_connections, "issues": issues}


def _detect_header(rows: list[tuple[Any, ...]]) -> tuple[int | None, dict[str, int]]:
    best_index: int | None = None
    best_score = 0
    best_map: dict[str, int] = {}
    for index, row in enumerate(rows[:25]):
        field_map: dict[str, int] = {}
        score = 0
        for col, value in enumerate(row):
            normalized = normalize_header(value)
            if not normalized:
                continue
            for field, aliases in HEADER_ALIASES.items():
                if normalized in aliases or any(alias in normalized for alias in aliases):
                    field_map.setdefault(field, col)
                    score += 2
            if " to " in normalized:
                score += 1
        if score > best_score:
            best_index = index
            best_score = score
            best_map = field_map
    if best_score < 2:
        return None, {}
    return best_index, best_map


def _extract_header_connections(
    sheet_name: str, headers: list[str], row: tuple[Any, ...], device_name: str, row_number: int
) -> list[dict[str, Any]]:
    connections: list[dict[str, Any]] = []
    for col, header in enumerate(headers):
        header_text = normalize_header(header)
        value = _get(row, col)
        if not value or " to " not in header_text:
            continue

        local_port, peer_hint = header_text.split(" to ", 1)
        peer_hint = peer_hint.strip()
        local_port = local_port.strip()

        if "pdu" in peer_hint:
            target_device = _device_alias(peer_hint)
            target_port = value
            cable_type = "power"
        elif peer_hint in {"sw p", "sw p#"} or "sw" in peer_hint and "#" in peer_hint:
            parsed_peer, parsed_port = _parse_peer_port(value)
            target_device = parsed_peer or "SW-Unknown"
            target_port = parsed_port
            cable_type = "management" if "idrac" in local_port or "mgmt" in local_port else "ethernet"
        elif any(token in peer_hint for token in ["sw", "switch", "fw", "firewall", "isp"]):
            if _parse_connection_text(value):
                continue
            target_device = _device_alias(peer_hint)
            target_port = value
            cable_type = "management" if "idrac" in local_port or "mgmt" in local_port else "ethernet"
        else:
            continue

        connections.append(
            {
                "sourceDevice": device_name,
                "sourcePort": local_port,
                "targetDevice": target_device,
                "targetPort": target_port,
                "cableType": cable_type,
                "role": _infer_role(local_port, cable_type),
                "source": f"{sheet_name}:{row_number}",
            }
        )
    return connections


def _scan_freeform_connections(
    sheet_name: str, rows: list[tuple[Any, ...]], starting_row: int = 1
) -> list[dict[str, Any]]:
    connections: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows, start=starting_row):
        for value in row:
            text = str(value or "").replace("\xa0", " ").strip()
            if not text:
                continue
            parsed = _parse_connection_text(text)
            if parsed:
                parsed["source"] = f"{sheet_name}:{row_index}"
                connections.append(parsed)
    return connections


def _parse_connection_text(text: str) -> dict[str, Any] | None:
    compact = re.sub(r"\s+", " ", text.strip(), flags=re.I)
    pattern = re.compile(
        r"(?P<src>swi?tch\s*\d+|sw\s*\d+|sw\d+|fw\s*\d+|fw\d+|isp\s*\d+|isp\d+)"
        r"\s*[- ]*(?:port)?\s*(?P<srcport>[A-Za-z0-9:/.]+)?\s*[- ]*to[- ]*"
        r"(?P<dst>swi?tch\s*\d+|sw\s*\d+|sw\d+|fw\s*\d+|fw\d+|firewall\s*\d+|isp\s*\d+|isp\d+)"
        r"\s*[- ]*(?:port)?\s*(?P<dstport>[A-Za-z0-9:/.]+)?",
        re.I,
    )
    match = pattern.search(compact)
    if match:
        return {
            "sourceDevice": _device_alias(match.group("src")),
            "sourcePort": match.group("srcport") or "",
            "targetDevice": _device_alias(match.group("dst")),
            "targetPort": match.group("dstport") or "",
            "cableType": "ethernet",
            "role": "lan",
        }

    isp_peer = re.search(r"(?P<peer>swi?tch\s*\d+|sw\s*\d+|sw\d+)\s*(?:port)?\s*(?P<port>[A-Za-z0-9:/.-]+)", compact, re.I)
    if isp_peer and "isp" in compact.lower():
        return {
            "sourceDevice": _device_alias(compact.split()[0]),
            "sourcePort": "",
            "targetDevice": _device_alias(isp_peer.group("peer")),
            "targetPort": isp_peer.group("port"),
            "cableType": "wan",
            "role": "wan",
        }
    return None


def _parse_peer_port(value: str) -> tuple[str | None, str | None]:
    text = str(value or "").replace("\xa0", " ").strip()
    match = re.match(r"(?P<peer>[A-Za-z]+-?\s*\d+)\s*(?:port)?\s*(?P<port>[A-Za-z0-9:/.-]+)?", text, re.I)
    if not match:
        return None, text or None
    return _device_alias(match.group("peer")), match.group("port")


def _device_alias(value: str) -> str:
    text = normalize_header(value)
    compact = text.replace(" ", "")
    match = re.match(r"(?:swi?tch|sw)(\d+)", compact)
    if match:
        return f"SW{match.group(1)}"
    match = re.match(r"(?:firewall|fw)(\d+)", compact)
    if match:
        return f"Firewall-{match.group(1)}"
    match = re.match(r"isp(\d+)", compact)
    if match:
        return f"ISP-{match.group(1)}"
    match = re.match(r"pdu(\d+)", compact)
    if match:
        return f"PDU-{match.group(1)}"
    return str(value or "").strip()


def _infer_role(port_or_role: str, cable_type: str) -> str:
    text = f"{port_or_role} {cable_type}".lower()
    if "idrac" in text or "ipmi" in text or "mgmt" in text:
        return "management"
    if "power" in text or "pdu" in text:
        return "power"
    if "wan" in text or "isp" in text:
        return "wan"
    return cable_type or "unknown"


def _get(row: tuple[Any, ...], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    value = row[index]
    return str(value).replace("\xa0", " ").strip() if value is not None else ""


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-") or "sheet"
